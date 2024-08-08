import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
input_file = "E:\\Projekty\\for_work\\ai-solution\\ITSquad-antyfraud\\excel\\dane-party-incident-user.xlsx"
wb = load_workbook(filename=input_file, data_only=True)
sheet = wb['MOCK_DATA']

# Function to convert Excel-style column letters to indices
def excel_column_index(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter) - 1


row_index = 4  
column_letters = ['AA', 'Y', 'Z', 'X', 'AN']  


column_indices = [excel_column_index(letter) for letter in column_letters]


row_indices = range(4, 787)  # For example, from row 4 to row 6

# Create a list to store the new data
new_data = []

for row_index in row_indices:
    # Extract data from the specified row and columns
    row_data = [sheet.cell(row=row_index, column=index + 1).value for index in column_indices]
    
    # Append each extracted value into new_data list as a dictionary
    for value in row_data:
        if isinstance(value, (int, float)):  # Check if value is numeric
            new_data.append({'W': str(value)})  # Convert numeric values to string before appending
        else:
            new_data.append({'W': value})

# Convert the list of dictionaries to a DataFrame
new_df = pd.DataFrame(new_data)

# Save the DataFrame to a new Excel file
output_file = 'output_party.xlsx'
new_df.to_excel(output_file, sheet_name='NewSheet', index=False)

print(f"Data has been successfully transformed and saved to {output_file}")
